"""Han 2025 (Nat Commun, s41467-025-64010-z) LbCas12a DR 变体数据提取与特征计算。

数据来源(全部已下载入库):
  MOESM3 XLSX: 工具箱 DR 序列(gfp-1 系 7 条 + P1 系 L3/L4)
  MOESM7 XLSX Fig.1 列63-64: Fig1g 活性表(归一化 GFP, 低=抑制强,
    2026-09-02 修正: 含 FL 组共 145 条, 旧版漏 48 条 FL)
  MOESM7 XLSX Fig.3b: 7 条工具箱两档表达(RBS0/RBS33)活性(n=3)
  data/han2025_sanger_sequences.json: MOESM1 Sup Fig.2/3 的 54 条
    Sanger 序列(600dpi 视觉转录+模板校验, 含置信度标注)
本脚本:
  1. 解析 MOESM3 提取 9 条工具箱 DR 序列(含 L3/L4);
  2. 对每条跑本管线特征(DDR-alone 折叠/ΔΔG/bp_dist/cross_pairs/stem_intact_prob/系综);
  3. Fig1g 145 条活性全表 + 工具箱 7 条 Fig1g 同尺度锚点值;
  4. 54 条 Sanger 耐受集特征计算(耐受先验, 无逐条活性标签);
  5. 与 Fig.3b 活性做 Spearman 相关(参考);
  6. (2026-09-03) 主文 Fig.1f 序列表视觉转录: 从编号-序列桥中恢复
     9 条新监督对(见 FIG1F_TRANSCRIPTION 溯源), 干净监督对 7 -> 16。
  7. (2026-09-03) Tian 2025 RRS 单点扫描入库(parse_tian2025):
     8 crRNA x 12 突变 = 96 条 LbCas12a trans-切割相对活性监督对。
  8. (2026-09-03) DeWeirdt 2020 AsCas12a 替代 DR 扫描入库(parse_deweirdt2020):
     35,883 条 20nt DR 变体 x 2 方向构建的负筛 LFC(全量特征写 data/raw/)。
诚实边界: LbCas12a CRISPRi(GFP 抑制)体系, 非 Cas12a2 RNA 触发杀伤;
  Fig1g 的 145 个编号(F/S/L/FL)与序列的映射未随论文数值源数据发表
  (2026-09-02 查证 MOESM1-7; 2026-09-03 复查 MOESM7 全部 25 个工作表 +
  MOESM3/MOESM4 全文, 数值侧确无映射)。主文 Fig.1f 以图像形式给出部分
  编号-序列对照(典型突变体表), 经视觉转录+双重独立读取+Sanger 交叉校验
  恢复 9 条新监督对(置信度分级见 fig1f_pairs); 其余 129 个编号仍不可得,
  完整映射需向作者索取。
运行: python scripts/crrna_han2025_features.py
"""
import argparse
import json
import os
import sys

import numpy as np
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, ".."))
sys.path.insert(0, HERE)

import crrna_scaffold_design as core  # noqa: E402

DATA = os.path.join(ROOT, "data")

# Fig.3b 工具箱活性(均值, 低=抑制强, 来自源数据手动提取)
TOOLBOX_ACTIVITY = {
    "CN": {"rbs0": 36.5, "rbs33": 50.9},
    "F1": {"rbs0": 29.6, "rbs33": 51.9},
    "F2": {"rbs0": 36.0, "rbs33": 59.5},
    "FL1": {"rbs0": 82.2, "rbs33": 92.3},
    "FL2": {"rbs0": 79.3, "rbs33": 87.1},
    "L1": {"rbs0": 82.8, "rbs33": 98.6},
    "L2": {"rbs0": 38.5, "rbs33": 67.8},
}

# 主文 Fig.1f 序列表视觉转录(2026-09-03)。
# 溯源: nature.com 文章页 Fig1 全图 PNG(2017x2201) -> panel f 右侧编号表
#   区域裁切(scripts/tmp/fig1f/, gitignore) -> 分块 2x-3x 放大 -> 视觉模型
#   两次独立逐字母转录 + 与 han2025_sanger_sequences.json(MOESM1 Sup Fig.2/3
#   独立转录源)的 loop 六联体(11-16 位)交叉校验。
# 置信度: high = 两次完整独立读取一致 + loop 六联体与 Sanger 独立源吻合;
#   medium = 单次完整读取 + loop 六联体与 Sanger 独立源吻合。
# 边界: 其余编号(含 FL3/FL5, 不在 Fig1g 活性表内)未收录; 视觉转录存在
#   单字母误差风险(如 S16 七连 C), 训练使用时建议按置信度加权或敏感性剔除。
FIG1F_TRANSCRIPTION = {
    "S3":  {"dr_rna": "UAAUUUCUACUCAUAGUAGAU", "confidence": "high",
            "cross_check": "loop UCAUAG = Sanger B19"},
    "S5":  {"dr_rna": "UAAUUUCUACUGUUUGUAGAU", "confidence": "high",
            "cross_check": "loop UGUUUG = Sanger A36/B19b"},
    "S7":  {"dr_rna": "UAAUUUCUACCCUAAGUAGAU", "confidence": "medium",
            "cross_check": "loop CCUAAG = Sanger B24(二读截边, 单完整读)"},
    "S10": {"dr_rna": "UAAUUUCUACCCGAGGUAGAU", "confidence": "medium",
            "cross_check": "loop CCGAGG = Sanger C48/C37"},
    "S14": {"dr_rna": "UAAUUUCUACGGGCGGUAGAU", "confidence": "medium",
            "cross_check": "loop GGGCGG = Sanger B58/D47"},
    "S16": {"dr_rna": "UAAUUUCUACCCCCCGUAGAU", "confidence": "medium",
            "cross_check": "loop CCCCCG = Sanger F33(完全一致, 同聚段易误读已复核)"},
    "S20": {"dr_rna": "UAAUUUCUACCAAUGGUAGAU", "confidence": "medium",
            "cross_check": "loop CAAUGG = Sanger B29"},
    "S21": {"dr_rna": "UAAUUUCUACUGGAUGUAGAU", "confidence": "medium",
            "cross_check": "loop UGGAUG = Sanger A40"},
    "FL4": {"dr_rna": "CAAUGUCUACCUUCUGUAGAU", "confidence": "medium",
            "cross_check": "flank CAAUGU+loop CUUCUG = Sanger B38"},
}


def parse_sequences():
    wb = openpyxl.load_workbook(
        os.path.join(DATA, "41467_2025_64010_MOESM3_ESM.xlsx"), data_only=True)
    ws = wb["Sheet1"]
    seqs = {}
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        seq = ws.cell(r, 2).value
        if not name or not seq:
            continue
        name = str(name).strip()
        seq = str(seq).replace(" ", "").replace("T", "U").upper()
        # gfp-1 系列(同一 spacer, 不同 DR)保证单变量;
        # L3/L4 只在 P1 系列出(取 DR 21nt, 同一 DR 跨系列一致)
        for prefix in ("gfp-1-", "P1-"):
            if name.startswith(prefix):
                tag = name.replace(prefix, "")
                if tag not in seqs:
                    seqs[tag] = seq
    return seqs


def extract_screen():
    """Fig.1 列63-64 的 Fig1g 活性表(含 FL 组, 共 145 条 + Canonical)"""
    import re
    wb = openpyxl.load_workbook(
        os.path.join(DATA, "41467_2025_64010_MOESM7_ESM.xlsx"),
        data_only=True, read_only=True)
    ws = wb["Fig. 1"]
    pat = re.compile(r"^(FL|F|S|L)\d+$")
    seen = {}
    for row in ws.iter_rows(min_row=1, min_col=63, max_col=64,
                            values_only=True):
        s, v = row
        if s is None or v is None:
            continue
        s = str(s).strip()
        if s == "Canonical" or pat.match(s):
            if s not in seen and isinstance(v, (int, float)):
                seen[s] = float(v)
    return seen


def extract_cleavage(sheet="Sup Fig. 5", col0=1):
    """Sup Fig.5(cis)/6(trans, b 面板) 实时切割曲线终点/起点比(三重复均值)。

    返回 {标签: 终点比}; 标签 Canonical -> CN。b 面板列布局: 7 个 crRNA x 3 重复。
    """
    wb = openpyxl.load_workbook(
        os.path.join(DATA, "41467_2025_64010_MOESM7_ESM.xlsx"),
        data_only=True, read_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c) if c is not None else "" for c in rows[1]]
    out = {}
    for i, h in enumerate(hdr):
        if h in ("Canonical", "L1", "L2", "FL1", "FL2", "F1", "F2"):
            tri = [j for j in range(i, min(i + 3, len(hdr)))
                   if rows[2][j] is not None]
            first = np.mean([float(rows[2][j]) for j in tri])
            last = np.mean([float(rows[-1][j]) for j in tri])
            tag = "CN" if h == "Canonical" else h
            out[tag] = round(float(last / max(first, 1e-9)), 3)
    return out


# === DeWeirdt 2020 alt-DR 扫描(2026-09-03 入库) ===
# 来源: DeWeirdt et al. 2020, Nat Biotechnol 39:94-104, doi:10.1038/s41587-020-0600-6
#   Supp Data 4 (= springer 41587_2020_600_MOESM6_ESM.xlsx, sheet alt_DR_reads):
#   35,682 随机化变体(茎区 ≤3 可变碱基对 + 单链/环区 ≤3 可变核苷酸)
#   + intend-6T/random 对照 + 1 条 wildtype, 共 35,883 行, 20nt DR 窗口。
# 活性口径: MELJUSO + 2xNLS-AsCas12a 合成致死负筛(BCL2L1xMCL1 双敲致死),
#   LFC = log2((Rep+1)/(pDNA+1)) 两重复均值, 低=活性强(功能 DR 致细胞耗竭);
#   两方向构建分别给读数: pRDA_127 = BCL2L1 guide-DR库-MCL1 guide(变体 DR 配 MCL1 spacer),
#   pRDA_128 = MCL1 guide-DR库-BCL2L1 guide(变体 DR 配 BCL2L1 spacer)。
# spacer 序列: MOESM1 Supplementary Table 3 "Individual guides used in this study"(Fig.4a)。
DEWEIRDT_WT_DR_DNA = "TAATTTCTACTCTTGTAGAT"  # alt_DR_reads type='wildtype' 行
DEWEIRDT_SPACERS = {  # DNA, 23nt; 键名 = 构建(变体 DR 的成熟 crRNA 下游 spacer)
    "pRDA_127": "ATGTCCAGTTTCCGAAGCATGCC",  # MCL1 guide (Fig.4a)
    "pRDA_128": "TGGGAAAGCTTGTAGGAGAGAAA",  # BCL2L1 guide (Fig.4a)
}


def parse_deweirdt2020():
    """DeWeirdt 2020 AsCas12a 替代 DR 扫描提取(2026-09-03)。

    返回 rows: 每变体 {tag, dr_rna, n_mut, cls, lfc127, lfc128,
      features_127, features_128}; 特征以各自方向 spacer 上下文、
      同方向 WT 为基准(ddG_dr/bp_dist), 口径与工具箱 compute_features 一致。
    源文件缺失返回 [] 并提示(不阻断)。
    """
    path = os.path.join(DATA, "41587_2020_600_MOESM6_ESM.xlsx")
    if not os.path.exists(path):
        print("DeWeirdt2020 源文件未入库(data/41587_2020_600_MOESM6_ESM.xlsx), 跳过")
        return []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["alt_DR_reads"]
    raw = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            counts = [float(c) for c in row[2:8]]
        except (TypeError, ValueError):
            continue
        raw.append((str(row[0]).strip().upper(), str(row[1] or "").strip(),
                    counts))
    wb.close()
    wt = [r for r in raw if r[1] == "wildtype"]
    if len(wt) != 1 or wt[0][0] != DEWEIRDT_WT_DR_DNA:
        raise SystemExit("DeWeirdt2020 WT 锚定失败: wildtype 行=%s" % (wt[:2]))
    # 各方向 WT 折叠基准
    base = {}
    for ori, sp_dna in DEWEIRDT_SPACERS.items():
        sp = sp_dna.replace("T", "U")
        wt_dr = DEWEIRDT_WT_DR_DNA.replace("T", "U")
        wt_full_ss, _ = core.fold(wt_dr + sp)
        _, wt_dr_mfe = core.fold(wt_dr)
        base[ori] = (sp, wt_dr, wt_full_ss, wt_dr_mfe)
    rows = []
    for seq_dna, cls, (a1, b1, p1, a2, b2, p2) in raw:
        dr = seq_dna.replace("T", "U")
        if len(dr) != 20 or set(dr) - set("AUCG"):
            continue
        entry = {"tag": "%s_%s" % (cls, seq_dna), "dr_rna": dr,
                 "cls": cls,
                 "n_mut": sum(a != b for a, b in zip(DEWEIRDT_WT_DR_DNA, seq_dna)),
                 "lfc127": round(float(np.mean([
                     np.log2((a1 + 1) / (p1 + 1)),
                     np.log2((b1 + 1) / (p1 + 1))])), 4),
                 "lfc128": round(float(np.mean([
                     np.log2((a2 + 1) / (p2 + 1)),
                     np.log2((b2 + 1) / (p2 + 1))])), 4)}
        for ori, key in (("pRDA_127", "127"), ("pRDA_128", "128")):
            sp, wt_dr, wt_full_ss, wt_dr_mfe = base[ori]
            feats = compute_features(dr, sp, wt_dr_mfe)
            v_ss, _ = core.fold(dr + sp)
            feats["bp_dist"] = RNA_bp_distance(wt_full_ss, v_ss)
            entry["features_" + key] = feats
        rows.append(entry)
    print("DeWeirdt2020 alt-DR 配对 %d 条(WT LFC: 127=%.3f, 128=%.3f)" % (
        len(rows),
        [r["lfc127"] for r in rows if r["cls"] == "wildtype"][0],
        [r["lfc128"] for r in rows if r["cls"] == "wildtype"][0]))
    return rows


def parse_tian2025():
    """Tian 2025 (Nat Commun, s41467-62082-5) LbCas12a RRS 单点扫描提取(2026-09-03)。

    序列: data/41467_2025_62082_MOESM3_ESM.xlsx Sheet1
      (name='WT crRNA-N' / '1A>U crRNA-N'; 全长 40nt = DR 20nt + spacer 20nt);
    活性: data/41467_2025_62082_MOESM6_ESM.xlsx Fig.1d
      (8 条 crRNA x (WT+12 单点突变) 的 trans-切割 ΔF, 三重复+均值列)。
    活性口径 rrs_rel = 同一条 crRNA 内 mut_mean/WT_mean(比值去除 spacer 基线, 高=活性强);
    特征以各 crRNA 自身 spacer 上下文计算、以各自 WT 为基准(ddG_dr/bp_dist 相对自身 WT)。
    返回 pairs 列表; 源文件缺失时返回 [] 并提示(不阻断)。
    """
    import re
    seq_path = os.path.join(DATA, "41467_2025_62082_MOESM3_ESM.xlsx")
    act_path = os.path.join(DATA, "41467_2025_62082_MOESM6_ESM.xlsx")
    if not (os.path.exists(seq_path) and os.path.exists(act_path)):
        print("Tian2025 源文件未入库(data/41467_2025_62082_MOESM3/6), 跳过")
        return []
    wt_seqs, mut_seqs = {}, {}
    wb = openpyxl.load_workbook(seq_path, data_only=True, read_only=True)
    for row in wb["Sheet1"].iter_rows(values_only=True):
        if not row or row[0] is None or len(row) < 2 or row[1] is None:
            continue
        name = str(row[0]).strip()
        seq = str(row[1]).replace(" ", "").replace("T", "U").upper()
        if name.startswith("WT crRNA-"):
            wt_seqs[int(name.rsplit("-", 1)[1])] = seq
            continue
        m = re.match(r"^(\d)([ACGU])>([ACGU]) crRNA-(\d)$", name)
        if m:
            pos, b_from, b_to, n = m.groups()
            mut_seqs[(f"{pos}{b_to}", int(n))] = seq
    wb2 = openpyxl.load_workbook(act_path, data_only=True, read_only=True)
    ws = wb2["Fig.1d"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    mean_cols = {}
    for j, h in enumerate(hdr):
        if h == "WT-mean":
            mean_cols["WT"] = j
        elif re.match(r"^\d[ACGU]-mean$", h):
            mean_cols[h[:-5]] = j
    acts = {}
    for r in rows[1:]:
        if not r or r[0] is None or not str(r[0]).startswith("crRNA-"):
            continue
        n = int(str(r[0]).split("-", 1)[1])
        for key, j in mean_cols.items():
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)):
                acts[(key, n)] = float(v)
    pairs = []
    for n in sorted({k[1] for k in acts}):
        wt_seq = wt_seqs.get(n)
        wt_act = acts.get(("WT", n))
        if not wt_seq or not wt_act:
            continue
        wt_dr, spacer = wt_seq[:20], wt_seq[20:40]
        wt_full_ss, _ = core.fold(wt_dr + spacer)
        _, wt_dr_mfe = core.fold(wt_dr)
        for key in ("1U", "1G", "1C", "2U", "2G", "2C",
                    "3A", "3G", "3C", "4A", "4G", "4C"):
            seq = mut_seqs.get((key, n))
            act = acts.get((key, n))
            if seq is None or act is None:
                continue
            dr = seq[:20]
            pos = int(key[0]) - 1
            if dr[pos] != key[1] or sum(a != b for a, b in zip(wt_dr, dr)) != 1:
                print("Tian2025 命名-序列不一致, 跳过: crRNA-%d %s" % (n, key))
                continue
            feats = compute_features(dr, spacer, wt_dr_mfe)
            v_ss, _ = core.fold(dr + spacer)
            feats["bp_dist"] = RNA_bp_distance(wt_full_ss, v_ss)
            pairs.append({"tag": "%s_crRNA%d" % (key, n), "crRNA": n,
                          "dr_rna": dr, "spacer_rna": spacer, "n_mut": 1,
                          "rrs_rel": round(act / wt_act, 4),
                          "trans_df_wt": round(wt_act, 1),
                          "trans_df_mut": round(act, 1), **feats})
    print("Tian2025 RRS 配对 %d 条(8 条 crRNA x 12 单点突变)" % len(pairs))
    return pairs


def compute_features(dr_rna, spacer_rna, wt_dr_mfe=0.0):
    full = dr_rna + spacer_rna
    ss, mfe = core.fold(full)
    dr_ss, dr_mfe = core.fold(dr_rna)
    cross_nt, cross_pp = core.cross_pairs(full, len(dr_rna))
    inv_run = core.inv_max_run(full, len(dr_rna))
    _, spacer_up, seed_up = core.pf_stats(full, len(dr_rna), 7)
    stem_pairs = core.stem_pairs_of(dr_ss)
    p_fold = core.stem_intact_prob(full, stem_pairs) if stem_pairs else 1.0
    return {
        "mfe": round(mfe, 2), "ddG_dr": round(dr_mfe - wt_dr_mfe, 2),
        "bp_dist": None,  # 需 WT 基准, 主循环里算
        "cross_nt": cross_nt, "cross_pp": round(cross_pp, 4),
        "inv_max_run": inv_run,
        "spacer_up": round(spacer_up, 3), "seed_up": round(seed_up, 3),
        "p_fold": round(p_fold, 5),
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--spacer-len", type=int, default=20,
                    help="spacer 用于折叠的截取长度(默认 20)")
    args = ap.parse_args()

    seqs = parse_sequences()
    if "CN" not in seqs:
        raise SystemExit("MOESM3 解析失败: 未找到 gfp-1-CN")
    spacer = seqs["CN"][21:21 + args.spacer_len]  # DR 21nt 后取 spacer
    wt_dr = seqs["CN"][:21]
    print("WT DR(21nt): %s" % wt_dr)
    print("Spacer(前%dnt): %s" % (len(spacer), spacer))

    wt_dr_ss, wt_dr_mfe = core.fold(wt_dr)
    wt_full_ss, wt_mfe = core.fold(wt_dr + spacer)
    print("WT DR MFE=%.2f, 全长 MFE=%.2f" % (wt_dr_mfe, wt_mfe))

    rows = []
    for tag, full_seq in sorted(seqs.items()):
        dr = full_seq[:21]
        feats = compute_features(dr, spacer, wt_dr_mfe)
        feats["bp_dist"] = None  # 后填
        act = TOOLBOX_ACTIVITY.get(tag, {})
        rows.append({"tag": tag, "dr_rna": dr, "n_mut": sum(
            1 for a, b in zip(wt_dr, dr) if a != b), **feats, **act})

    # bp_dist vs WT
    for r in rows:
        v_ss, _ = core.fold(r["dr_rna"] + spacer)
        r["bp_dist"] = RNA_bp_distance(wt_full_ss, v_ss)

    print("\n%-6s %4s %8s %8s %6s %6s %7s %7s %8s %8s" % (
        "var", "nmut", "ddG_dr", "MFE", "bp_d", "x_nt", "P_fold", "sp_up",
        "RBS0", "RBS33"))
    for r in rows:
        print("%-6s %4d %8.2f %8.2f %6d %6d %7.4f %7.3f %8.1f %8.1f" % (
            r["tag"], r["n_mut"], r["ddG_dr"], r["mfe"], r["bp_dist"],
            r["cross_nt"], r["p_fold"], r["spacer_up"],
            r.get("rbs0", 0), r.get("rbs33", 0)))

    # Spearman 相关(仅 7 点, 方向参考)
    acts0 = np.array([r.get("rbs0", np.nan) for r in rows])
    acts33 = np.array([r.get("rbs33", np.nan) for r in rows])
    for feat in ("ddG_dr", "bp_dist", "cross_nt", "p_fold", "spacer_up"):
        vals = np.array([float(r[feat]) for r in rows])
        mask = ~np.isnan(acts0)
        rho0 = spearman(vals[mask], acts0[mask])
        rho33 = spearman(vals[mask], acts33[mask])
        print("%-10s  rho(RBS0)=%+.3f  rho(RBS33)=%+.3f" % (feat, rho0, rho33))

    # === Fig1g 活性全表 + 工具箱同尺度锚点(程序化提取, 不硬编码) ===
    fig1g = extract_screen()
    tag_map = {"CN": "Canonical"}
    anchors_fig1g = {}
    for r in rows:
        key = tag_map.get(r["tag"], r["tag"])
        if key in fig1g:
            r["fig1g"] = fig1g[key]
            anchors_fig1g[r["tag"]] = fig1g[key]
    print("\nFig1g 全表 %d 条; 工具箱同尺度锚点 %d 条: %s" % (
        len(fig1g), len(anchors_fig1g), anchors_fig1g))

    # === Fig.1f 视觉转录配对(2026-09-03): 编号->序列->Fig1g 活性 ===
    fig1f_pairs = []
    for tag, e in sorted(FIG1F_TRANSCRIPTION.items()):
        dr = e["dr_rna"]
        assert len(dr) == 21 and set(dr) <= set("AUCG"), tag
        feats = compute_features(dr, spacer, wt_dr_mfe)
        v_ss, _ = core.fold(dr + spacer)
        feats["bp_dist"] = RNA_bp_distance(wt_full_ss, v_ss)
        fig1f_pairs.append({
            "tag": tag, "dr_rna": dr, "confidence": e["confidence"],
            "cross_check": e["cross_check"],
            "n_mut": sum(1 for a, b in zip(wt_dr, dr) if a != b),
            "fig1g": fig1g.get(tag), **feats})
    new_tags = [p["tag"] for p in fig1f_pairs if p["fig1g"] is not None]
    print("Fig1f 转录配对 %d 条, 其中带 Fig1g 活性 %d 条: %s" % (
        len(fig1f_pairs), len(new_tags), new_tags))

    # === 同源多终点: cis(Sup Fig.5b) / trans(Sup Fig.6b) 切割终点比 ===
    cis = extract_cleavage("Sup Fig. 5")
    trans = extract_cleavage("Sup Fig. 6")
    for r in rows:
        if r["tag"] in cis:
            r["cis_end"] = cis[r["tag"]]
        if r["tag"] in trans:
            r["trans_end"] = trans[r["tag"]]
    print("cis 终点比: %s" % cis)
    print("trans 终点比: %s" % trans)

    # === Sanger 耐受集(MOESM1 Sup Fig.2/3, 54 条)特征计算 ===
    sanger_path = os.path.join(DATA, "han2025_sanger_sequences.json")
    sanger_rows = []
    if os.path.exists(sanger_path):
        sanger = json.load(open(sanger_path, encoding="utf-8"))
        for e in sanger["entries"]:
            dr = e["seq"]
            feats = compute_features(dr, spacer, wt_dr_mfe)
            v_ss, _ = core.fold(dr + spacer)
            feats["bp_dist"] = RNA_bp_distance(wt_full_ss, v_ss)
            sanger_rows.append({
                "label": e["label"], "group": e["group"],
                "dr_rna": dr, "confidence": e["confidence"],
                "reads": e.get("reads", 1),
                "matches_toolbox": e.get("matches_toolbox"),
                "n_mut": sum(1 for a, b in zip(wt_dr, dr) if a != b),
                **feats})
        print("Sanger 耐受集 %d 条特征计算完成" % len(sanger_rows))

    # === Tian 2025 RRS 单点扫描(2026-09-03 入库): 8 crRNA x 12 突变 ===
    tian_pairs = parse_tian2025()

    # === DeWeirdt 2020 AsCas12a 替代 DR 扫描(2026-09-03 入库): 35,883 条 ===
    # 全量特征表较大, 写 data/raw/(不入 git, 可由本脚本重算); 溯源入 dataset json
    dw_rows = parse_deweirdt2020()
    if dw_rows:
        dw_path = os.path.join(DATA, "raw", "deweirdt2020_dr_scan.json")
        os.makedirs(os.path.dirname(dw_path), exist_ok=True)
        json.dump({
            "source": "DeWeirdt et al. 2020 Nat Biotechnol s41587-020-0600-6 "
                      "(AsCas12a alt-DR 负筛); 序列+读数=MOESM6 alt_DR_reads; "
                      "spacer=MOESM1 Supp Table 3 (MCL1/BCL2L1); "
                      "lfc127/lfc128 = log2((Rep+1)/(pDNA+1)) 两重复均值, 低=活性强",
            "wt_dr_dna": DEWEIRDT_WT_DR_DNA, "spacers": DEWEIRDT_SPACERS,
            "n": len(dw_rows), "rows": dw_rows},
            open(dw_path, "w", encoding="utf-8"), ensure_ascii=False)
        print("DeWeirdt2020 全量特征 -> %s (n=%d)" % (dw_path, len(dw_rows)))

    out = {"toolbox_sequences": {t: s for t, s in seqs.items()},
           "toolbox_features": rows,
           "toolbox_activity": TOOLBOX_ACTIVITY,
           "toolbox_activity_fig1g": anchors_fig1g,
           "toolbox_cleavage": {"cis_end": cis, "trans_end": trans},
           "fig1g_activity": fig1g,
           "fig1f_pairs": fig1f_pairs,
           "tian2025_rrs_pairs": tian_pairs,
           "deweirdt2020_source": (
               "DeWeirdt et al. 2020 Nat Biotechnol s41587-020-0600-6 (AsCas12a "
               "alt-DR 负筛 n=%d); 全量特征表在 data/raw/deweirdt2020_dr_scan.json "
               "(不入 git); LFC 低=活性强; 特征按方向 spacer 上下文+同向 WT 基准"
               % len(dw_rows)),
           "tian2025_source": (
               "Tian et al. 2025 Nat Commun s41467-025-62082-5 (LbCas12a RRS "
               "单点扫描); 序列=MOESM3, 活性=MOESM6 Fig.1d 三重复均值; "
               "rrs_rel = 同 crRNA 内 mut/WT ΔF 比值(高=活性强); 特征以各 crRNA "
               "自身 spacer 上下文计算、以各自 WT 为基准"),
           "fig1f_source": (
               "主文 Fig.1f 序列表视觉转录(2026-09-03): nature.com Fig1 全图 "
               "panel f 右侧编号表, 分块放大两次独立逐字母读取 + Sanger "
               "独立转录源 loop 六联体交叉校验; 置信度分级见各条 confidence; "
               "其余 129 个编号的序列映射仍不可得(数值源数据无, 需向作者索取)"),
           "sanger_tolerance": sanger_rows,
           "sanger_source": (sanger["source"] + "; " + sanger["extraction"]
                             if sanger_rows else None),
           "source": "Han et al. 2025 Nat Commun s41467-025-64010-z "
                     "(LbCas12a CRISPRi)"}
    # 合并写: 保留 rejected_endpoints 等后加键, 避免整体重写丢数据
    ds_path = os.path.join(DATA, "han2025_dataset.json")
    if os.path.exists(ds_path):
        try:
            old = json.load(open(ds_path, encoding="utf-8"))
            if isinstance(old, dict):
                old.update(out)
                out = old
        except (ValueError, OSError):
            pass
    json.dump(out, open(ds_path, "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    print("\n数据集 -> data/han2025_dataset.json (工具箱%d + Fig1g %d + "
          "Fig1f配对%d + Sanger耐受%d)" % (
              len(rows), len(fig1g), len(fig1f_pairs), len(sanger_rows)))


def RNA_bp_distance(s1, s2):
    import RNA
    return int(RNA.bp_distance(s1, s2))


def spearman(x, y):
    rx = np.argsort(np.argsort(x)).astype(float)
    ry = np.argsort(np.argsort(y)).astype(float)
    return float(np.corrcoef(rx, ry)[0, 1])


if __name__ == "__main__":
    main()
